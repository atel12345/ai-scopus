"""
Module 5 - Generateur Excel

Genere le fichier .xlsx final avec deux feuilles :
- Publications : tableau detaille (section 12 du cahier des charges)
- Statistiques : synthese avec formules (section 13)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelGenerationError(Exception):
    pass


COLUMNS = [
    "N°", "Paper Title", "Authors", "Publication Year", "First Author",
    "Journal", "ISSN", "DOI", "Quartile", "SCImago Category", "Scopus Link",
]


def write_publications_sheet(workbook, publications: list[dict]):
    """
    Cree et remplit la feuille 'Publications' avec les colonnes du
    cahier des charges (section 12).

    Args:
        workbook: objet openpyxl.Workbook actif
        publications: liste de dicts, chacun avec les cles attendues
                       (title, authors, year, first_author, journal,
                        issn, doi, quartile, scimago_category, scopus_link)
    """
    sheet = workbook.active
    sheet.title = "Publications"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, pub in enumerate(publications, start=2):
        sheet.cell(row=row_idx, column=1, value=row_idx - 1)
        sheet.cell(row=row_idx, column=2, value=pub.get("title", ""))
        sheet.cell(row=row_idx, column=3, value=pub.get("authors", ""))
        sheet.cell(row=row_idx, column=4, value=pub.get("year", ""))
        sheet.cell(row=row_idx, column=5, value=pub.get("first_author", ""))
        sheet.cell(row=row_idx, column=6, value=pub.get("journal", ""))
        sheet.cell(row=row_idx, column=7, value=pub.get("issn", ""))
        sheet.cell(row=row_idx, column=8, value=pub.get("doi", ""))
        sheet.cell(row=row_idx, column=9, value=pub.get("quartile", ""))
        sheet.cell(row=row_idx, column=10, value=pub.get("scimago_category", ""))
        sheet.cell(row=row_idx, column=11, value=pub.get("scopus_link", ""))

    for col_idx, header in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            [len(str(header))]
            + [len(str(sheet.cell(row=r, column=col_idx).value or "")) for r in range(2, len(publications) + 2)]
        )
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    for col_letter in ["A"]:
        sheet.column_dimensions[col_letter].width = 6

    sheet.freeze_panes = "A2"


def write_statistics_sheet(workbook, num_publications: int):
    """
    Cree la feuille 'Statistiques' avec des formules Excel qui se
    recalculent automatiquement a partir de la feuille Publications
    (section 13 du cahier des charges).

    Args:
        workbook: objet openpyxl.Workbook actif (Publications deja creee)
        num_publications: nombre de lignes de donnees dans Publications
                           (utilise pour bien delimiter les plages de formule)
    """
    sheet = workbook.create_sheet("Statistiques")

    last_row = num_publications + 1  # +1 car les donnees commencent ligne 2

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    label_font = Font(name="Arial", bold=False)

    sheet.cell(row=1, column=1, value="Indicateur").font = header_font
    sheet.cell(row=1, column=1).fill = header_fill
    sheet.cell(row=1, column=2, value="Valeur").font = header_font
    sheet.cell(row=1, column=2).fill = header_fill

    rows = [
        ("Publications totales", f"=COUNTA(Publications!A2:A{last_row})"),
        ("Q1", f'=COUNTIF(Publications!I2:I{last_row},"Q1")'),
        ("Q2", f'=COUNTIF(Publications!I2:I{last_row},"Q2")'),
        ("Q3", f'=COUNTIF(Publications!I2:I{last_row},"Q3")'),
        ("Q4", f'=COUNTIF(Publications!I2:I{last_row},"Q4")'),
        ("Quartile non disponible", f'=COUNTBLANK(Publications!I2:I{last_row})'),
        ("Premier auteur", f'=COUNTIF(Publications!E2:E{last_row},"OUI")'),
        ("Co-auteur", f'=COUNTIF(Publications!E2:E{last_row},"NON")'),
    ]

    for row_idx, (label, formula) in enumerate(rows, start=2):
        sheet.cell(row=row_idx, column=1, value=label).font = label_font
        sheet.cell(row=row_idx, column=2, value=formula)

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 15


def generate_excel_report(publications: list[dict], output_path: str):
    """
    Genere le fichier Excel final complet (Publications + Statistiques).

    Args:
        publications: liste de dicts representant chaque publication
        output_path: chemin de sortie du fichier .xlsx

    Raises:
        ExcelGenerationError si la generation echoue.
    """
    try:
        workbook = openpyxl.Workbook()
        write_publications_sheet(workbook, publications)
        write_statistics_sheet(workbook, len(publications))
        workbook.save(output_path)
    except Exception as e:
        raise ExcelGenerationError(f"Échec de la génération du fichier Excel : {e}") from e